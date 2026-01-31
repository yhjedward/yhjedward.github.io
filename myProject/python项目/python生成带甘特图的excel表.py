import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import datetime
import io

# =============================================
# 步骤1：准备项目进度数据
# =============================================
# 项目任务数据（任务名、开始日期、结束日期、完成百分比）
data = {
    "任务": [
        "需求分析", "系统设计", "前端开发", "后端开发", "测试阶段", "部署上线"
    ],
    "开始日期": [
        "2023-10-01", "2023-10-05", "2023-10-10", "2023-10-12", "2023-10-25", "2023-11-01"
    ],
    "结束日期": [
        "2023-10-04", "2023-10-09", "2023-10-24", "2023-10-26", "2023-10-31", "2023-11-05"
    ],
    "完成百分比": [100, 100, 70, 50, 30, 0]
}

# 转换为 DataFrame
df = pd.DataFrame(data)
df["开始日期"] = pd.to_datetime(df["开始日期"])
df["结束日期"] = pd.to_datetime(df["结束日期"])
df["持续天数"] = (df["结束日期"] - df["开始日期"]).dt.days + 1  # 计算持续天数

plt.rcParams['font.sans-serif'] = ['SimHei'] # 或者 'Microsoft YaHei', 'Noto Sans CJK'
plt.rcParams['axes.unicode_minus'] = False # 解决负号显示问题

# =============================================
# 步骤2：生成甘特图
# =============================================
plt.figure(figsize=(12, 8))  # 设置图形大小
plt.title("项目进度甘特图", fontsize=16)

# 为每个任务绘制水平条形
for idx, row in df.iterrows():
    start = row["开始日期"]
    end = row["结束日期"]
    duration = (end - start).days
    # 条形图的左边位置 = 开始日期的天数（相对于第一个任务）
    plt.barh(
        y=idx,  # y 轴坐标（任务索引）
        width=duration,  # 条形宽度（持续天数）
        left=start,  # 条形起始位置
        color=plt.cm.viridis(row["完成百分比"] / 100),  # 根据完成度设置颜色
        edgecolor='black',
        label=f"{row['任务']} ({row['完成百分比']}%)"
    )

# 设置 y 轴刻度（任务名称）
plt.yticks(ticks=range(len(df)), labels=df["任务"], fontsize=12)

# 设置 x 轴格式（显示日期）
plt.gca().xaxis_date()  # 识别 x 轴为日期
plt.gcf().autofmt_xdate()  # 斜体显示并避免重叠

# 添加网格
plt.grid(axis='x', linestyle='--', alpha=0.7)

# 移除 legend（避免重叠）
plt.legend().set_visible(False)

# 保存图表到内存
img_buffer = io.BytesIO()
plt.savefig(img_buffer, format='png', dpi=80)
plt.close()  # 关闭图形以释放内存

# =============================================
# 步骤3：创建 Excel 文件并插入数据和图表
# =============================================
# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "项目进度"

# =============================
# 写入表头
# =============================
headers = ["任务", "开始日期", "结束日期", "持续天数", "完成百分比"]
for col_idx, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_idx, value=header)

# =============================
# 写入任务数据
# =============================
for row_idx, row in df.iterrows():
    row_num = row_idx + 2  # 从第2行开始写入（第1行是表头）
    ws.cell(row=row_num, column=1, value=row["任务"])
    ws.cell(row=row_num, column=2, value=row["开始日期"].strftime("%Y-%m-%d"))
    ws.cell(row=row_num, column=3, value=row["结束日期"].strftime("%Y-%m-%d"))
    ws.cell(row=row_num, column=4, value=row["持续天数"])
    ws.cell(row=row_num, column=5, value=f"{row['完成百分比']}%")

# =============================
# 设置表格格式（可选）
# =============================
# 自动调整列宽
for col in range(1, 6):
    max_length = 0
    column = get_column_letter(col)
    for row in range(1, row_idx + 3):  # 包括表头和数据
        cell = ws.cell(row=row, column=col)
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
    ws.column_dimensions[column].width = max_length + 4

# =============================
# 插入甘特图
# =============================
# 加载图片并插入到Excel
img = Image(img_buffer)
img_buffer.seek(0)  # 重置缓冲区指针

# 设置图片位置（在表格下方）
insert_row = len(df) + 3  # 在数据行下方3行处插入
ws.add_image(img, f"D{insert_row}")  # 从D列开始插入

# =============================
# 添加图表标题（可选）
# =============================
title_cell = ws.cell(row=insert_row - 2, column=1, value="📊 项目进度甘特图")
title_cell.font = Font(name=title_cell.font.name, size=title_cell.font.size, bold=True)

# =============================
# 保存 Excel 文件
# =============================
file_path = "项目进度甘特图.xlsx"
wb.save(file_path)
print(f"✅ 成功生成 Excel 文件: {file_path}")